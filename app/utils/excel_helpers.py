# Copyright (c) 2005-2025 Nitro Software Inc. All Rights Reserved

"""Excel file creation helpers for PDF extraction results"""

from io import BytesIO
from typing import Any

import openpyxl as pxl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pydantic import BaseModel, ConfigDict
from pydantic.alias_generators import to_camel


class JsonSchema(BaseModel):
    """Base model with camelCase alias generator for parsing API JSON responses."""

    model_config = ConfigDict(alias_generator=to_camel, populate_by_name=True)


class FormField(JsonSchema):
    """A single extracted form field."""

    name: str
    value: str
    confidence: float


class FormsResult(JsonSchema):
    """Top-level response from the forms extraction endpoint."""

    fields: list[FormField]
    average_confidence: float


class TableData(JsonSchema):
    """Data for a single extracted table."""

    title: str | None
    cells: list[list[Any]]
    average_confidence: float


class ExtractedTable(JsonSchema):
    """A table entry in the tables extraction response."""

    table_data: TableData
    page_indices: list[int] = []


class TablesResult(JsonSchema):
    """Top-level response from the tables extraction endpoint."""

    tables: list[ExtractedTable]


def _auto_adjust_columns(ws: Worksheet, min_width: int, max_width: int) -> None:
    for column_cells in ws.columns:
        cells = list(column_cells)
        length = max(len(str(cell.value) if cell.value is not None else "") for cell in cells)
        col_letter = get_column_letter(cells[0].column or 1)
        ws.column_dimensions[col_letter].width = min(max(length, min_width), max_width)


def _workbook_to_bytes(workbook: pxl.Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def create_forms_excel(fields: list[FormField], filename: str, avg_confidence: float) -> bytes:
    """Create Excel bytes from extracted forms data."""
    workbook = pxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    ws: Worksheet = workbook.create_sheet("Form Data")
    ws.append(["Document", "Field Name", "Field Value", "Confidence"])

    for field in fields:
        if field.value in ("NOT_SELECTED", "None"):
            value = "Not filled"
        elif field.value == "SELECTED":
            value = "Selected"
        else:
            value = field.value
        ws.append([filename, field.name, value, f"{field.confidence:.1%}"])

    ws.append([
        filename,
        "SUMMARY",
        f"Total Fields: {len(fields)}",
        f"Avg Confidence: {avg_confidence:.1%}",
    ])

    _auto_adjust_columns(ws, min_width=15, max_width=50)
    return _workbook_to_bytes(workbook)


def create_tables_excel(  # pylint: disable=too-many-locals
    tables: list[ExtractedTable], filename: str
) -> bytes:
    """Create Excel bytes from extracted tables data."""
    workbook = pxl.Workbook()
    if "Sheet" in workbook.sheetnames:
        del workbook["Sheet"]

    contents_sheet = workbook.create_sheet("Contents")
    contents_sheet.append(["Sheet", "Document", "Table", "Title", "Confidence", "Pages"])

    for i, table in enumerate(tables, 1):
        title = table.table_data.title or ""
        page_indices = table.page_indices

        sheet_name = f"Table {i}" if len(f"Table {i}") <= 31 else f"T{i}"
        ws: Worksheet = workbook.create_sheet(sheet_name)

        for row in table.table_data.cells:
            ws.append(row)

        _auto_adjust_columns(ws, min_width=10, max_width=50)

        pages_text = (
            f"Pages {', '.join(str(p + 1) for p in page_indices)}" if page_indices else "Unknown"
        )
        contents_sheet.append([
            sheet_name,
            filename,
            f"Table {i}",
            title,
            f"{table.table_data.average_confidence:.1%}",
            pages_text,
        ])

        cell: Cell = contents_sheet.cell(row=i + 1, column=1)
        cell.hyperlink = f"#'{sheet_name}'!A1"
        cell.style = "Hyperlink"

    return _workbook_to_bytes(workbook)
